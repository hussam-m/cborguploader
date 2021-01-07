#!/usr/bin/env python

import click as ck
import openpyxl
import yaml


@ck.command()
@ck.option(
    '--input-file', '-i', default='metadata.xlsx',
    help='Metadata in excel sheet')
@ck.option(
    '--metadata-file', '-mf', default='example/metadata.yaml',
    help='The directory to output the metadata')
@ck.option(
    '--output-dir', '-o', default='metadata/',
    help='The directory to output the metadata')
def main(input_file, metadata_file, output_dir):
    wb = openpyxl.load_workbook(input_file) 
    sheet = wb.active 

    header = []
    for i in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=i)
        header.append(cell.value.strip(' :'))
    data = []
    for i in range(1, sheet.max_row):
        item = {}
        for j in range(sheet.max_column):
            item[header[j]] = sheet.cell(row=i+1, column=j+1).value.strip()
        data.append(item)

    min_metadata = yaml.load(open(metadata_file), Loader=yaml.FullLoader)
    for item in data:
        metadata = min_metadata.copy()
        metadata['host']['host_id'] = item['host_id']
        metadata['sample']['sample_id'] = item['sample_id']
        metadata['sample']['collection_date'] = item['collection_date']
        metadata['sample']['collection_location'] = 'http://www.wikidata.org/entity/' + item['collection_location']
        metadata['sample']['specimen_source'] = ['http://purl.obolibrary.org/obo/NCIT_' + item['specimen_source']]
        with open(output_dir + item['sample_id'] + '.yaml', 'w') as f:
            yaml.dump(metadata, f)

if __name__ == '__main__':
    main()
